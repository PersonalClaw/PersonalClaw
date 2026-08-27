"""SheetModel → .xlsx bytes (openpyxl).

Cell TYPES are preserved deliberately: a spreadsheet whose numbers arrived as text can't
be summed or charted, which defeats the purpose of generating one.

**A formula is written as a formula, and a literal is written as a literal — because the
model says which it is.** openpyxl decides a cell's type by SNIFFING the string it is
given: ``cell.value = "=SUM(A1)"`` becomes a formula, and so does
``cell.value = "=WORK IN PROGRESS"``, which Excel then opens as ``#NAME?``. Sniffing is
wrong in both directions, so this writer never lets it happen: a
:class:`~personalclaw.documents.model.SheetCell` with a ``formula`` is written through
the formula path, and one with only a ``value`` is pinned to the string type when it
would otherwise be mistaken for an expression.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING, Any

from personalclaw.documents.model import Sheet, SheetCell, SheetModel
from personalclaw.documents.registry import register_writer

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.worksheet.worksheet import Worksheet

#: Excel's own limit on a sheet name, plus the characters it forbids.
_MAX_SHEET_NAME = 31
_ILLEGAL_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel refuses some names outright; a rejected name would fail the whole write."""
    cleaned = "".join("-" if ch in _ILLEGAL_SHEET_CHARS else ch for ch in (name or "").strip())
    cleaned = (cleaned or "Sheet")[:_MAX_SHEET_NAME]
    base, n = cleaned, 2
    while cleaned.casefold() in used:  # Excel treats sheet names case-insensitively
        suffix = f"-{n}"
        cleaned = f"{base[: _MAX_SHEET_NAME - len(suffix)]}{suffix}"
        n += 1
    used.add(cleaned.casefold())
    return cleaned


def _column_letter(index: int) -> str:
    """0-based column index → ``A``, ``B`` … ``AA``. openpyxl keys widths by letter."""
    from openpyxl.utils import get_column_letter

    return str(get_column_letter(index + 1))


def render_xlsx(model: object) -> bytes:
    from openpyxl import Workbook

    if not isinstance(model, SheetModel):
        raise TypeError("xlsx writer expects a SheetModel")
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; every sheet here is explicit
    used: set[str] = set()
    for sheet in model.sheets:
        ws = wb.create_sheet(_safe_sheet_name(sheet.name, used))
        _write_sheet(ws, sheet)
    if not wb.sheetnames:  # a model with no sheets still has to produce a valid file
        wb.create_sheet("Sheet1")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_sheet(ws: Worksheet, sheet: Sheet) -> None:
    for row_index, row in enumerate(sheet.cells):
        for col_index, cell in enumerate(row):
            _write_cell(ws.cell(row=row_index + 1, column=col_index + 1), cell)
    for col_index, width in enumerate(sheet.column_widths):
        if width > 0:  # 0.0 means "the writer's default", so touch nothing
            ws.column_dimensions[_column_letter(col_index)].width = width
    for ref in sheet.merges:
        ws.merge_cells(ref)
    if sheet.frozen_header:
        ws.freeze_panes = "A2"  # header stays visible while scrolling


def _write_cell(target: Any, cell: SheetCell) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    if cell.formula:
        # The one place a leading "=" is MEANT to be sniffed — the model declared it.
        target.value = cell.formula
    else:
        target.value = _literal(cell.value)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            # openpyxl already typed this as a formula on the line above. Pin it back to
            # a string: the model said `value`, not `formula`, and a label a user typed
            # must not become #NAME? in their spreadsheet. `data_type` survives the save.
            target.data_type = "s"
    if cell.number_format:
        target.number_format = cell.number_format
    if cell.bold or cell.italic or cell.font_color:
        target.font = Font(
            bold=cell.bold or None,
            italic=cell.italic or None,
            color=cell.font_color or None,
        )
    if cell.fill:
        target.fill = PatternFill(fill_type="solid", start_color=cell.fill)
    if cell.align:
        target.alignment = Alignment(horizontal=cell.align)


def _literal(value: object) -> object:
    """Pass through the types openpyxl writes natively; stringify anything else.

    bool is checked before int on purpose — in Python `bool` IS an `int`, and writing
    True as 1 would lose the distinction the model went out of its way to preserve.
    """
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    return str(value)


register_writer("xlsx", render_xlsx)
