""".xlsx bytes → :class:`SheetModel` + a :class:`LossReport`.

The READ half of the spreadsheet round trip. Without it a sheet could be written but
never loaded, so the editing surface had nothing to edit and no formula could survive a
lap: ``documents/writers/xlsx_writer.py`` could put ``=SUM(A1)`` in a file, and reading
it back gave the *string* ``"=SUM(A1)"`` with no way to tell it from a label.

**A formula is recognised from the FILE, not from the text.** openpyxl reports a cell's
own ``data_type``, so ``"f"`` means the file said formula and ``"s"`` means the file said
string — even when both hold text starting with ``=``. That is the whole reason the
distinction survives, and it is why this parser never inspects the leading character.

**Every parse lives beside a loss report, same contract as the docx parser.** A
spreadsheet holds a great deal this model does not (charts, pivot tables, conditional
formatting, data validation, borders, per-row heights), and the editor re-renders from
the model — so anything not represented here is gone the moment a save lands. It is
reported, never dropped silently, and located in the sheet's own terms (``Sales!C2``).
"""

from __future__ import annotations

import io
from typing import Any

from personalclaw.documents.docx_parser import LossReport
from personalclaw.documents.model import ALIGNMENTS, Sheet, SheetCell, SheetModel

#: Alignments the model can hold. openpyxl also reports ``general``/``fill``/
#: ``centerContinuous``/``distributed``, which have no model field — those become a
#: ``cell_style`` loss rather than being coerced into a near-neighbour.
_MODEL_ALIGNMENTS = frozenset(a for a in ALIGNMENTS if a)

#: Sheet-level features with no model field, mapped to the attribute that reveals them.
#: A table rather than a chain of ``if``s so the report cannot fall behind the check.
_SHEET_FEATURES: tuple[tuple[str, str], ...] = (
    ("conditional formatting", "conditional_formatting"),
    ("data validation", "data_validations"),
    ("auto filter", "auto_filter"),
    ("charts", "_charts"),
    ("images", "_images"),
    ("pivot tables", "_pivots"),
    ("tables", "tables"),
)


def parse_xlsx(data: bytes) -> tuple[SheetModel, LossReport]:
    """Parse *data* into a :class:`SheetModel` plus everything that did not fit.

    ``data_only=False`` on purpose: the alternative hands back cached results and throws
    the formulas away, which would turn every re-render into a spreadsheet of frozen
    numbers. The cost is that the cached result is not available, and that cost is
    REPORTED (``formula_cached_value``) rather than hidden.
    """
    from openpyxl import load_workbook

    report = LossReport()
    wb = load_workbook(io.BytesIO(data), data_only=False)
    try:
        sheets = [_sheet(ws, report) for ws in wb.worksheets]
    finally:
        wb.close()
    if wb.defined_names:
        report.add(
            "sheet_feature",
            f"{len(wb.defined_names)} defined name(s) are not carried by the model",
            location="workbook",
        )
    return SheetModel(sheets=sheets), report


def _sheet(ws: Any, report: LossReport) -> Sheet:
    name = str(ws.title)
    merges = sorted(str(ref) for ref in ws.merged_cells.ranges)
    cells = [
        [_cell(cell, report, f"{name}!{cell.coordinate}") for cell in row]
        for row in ws.iter_rows(min_row=1, min_col=1)
    ]
    _report_sheet_features(ws, name, report)
    return Sheet(
        name=name,
        cells=cells,
        column_widths=_column_widths(ws),
        merges=merges,
        frozen_header=str(ws.freeze_panes or "") == "A2",
    )


def _report_sheet_features(ws: Any, name: str, report: LossReport) -> None:
    """Name the sheet-level constructs the model cannot hold, one item each."""
    for label, attribute in _SHEET_FEATURES:
        present = getattr(ws, attribute, None)
        if attribute == "conditional_formatting":
            count = len(list(present)) if present is not None else 0
        elif attribute == "auto_filter":
            count = 1 if present is not None and getattr(present, "ref", None) else 0
        else:
            count = len(present) if present else 0
        if count:
            report.add(
                "sheet_feature",
                f"{count} {label} on this sheet are not carried by the model",
                location=name,
            )
    if str(ws.freeze_panes or "") not in ("", "A2"):
        report.add(
            "sheet_feature",
            f"frozen panes at {ws.freeze_panes} — the model holds only a frozen header row",
            location=name,
        )
    for key, dim in sorted(ws.row_dimensions.items()):
        if dim.height:
            report.add(
                "row_height",
                f"row {key} sets an explicit height of {dim.height}",
                location=f"{name}!{key}",
            )


def _column_widths(ws: Any) -> list[float]:
    """Widths as a dense, index-aligned list — the model's shape, not openpyxl's dict.

    Dense because a sparse map keyed by letter would put A1-notation arithmetic in the
    model, and index-aligned so a writer can zip it against the columns it is emitting.
    Trailing zeros are trimmed: they carry no information and a list of 16,384 zeros for
    a two-column sheet would be most of the payload.
    """
    from openpyxl.utils import column_index_from_string

    widths: dict[int, float] = {}
    for letter, dim in ws.column_dimensions.items():
        if dim.width and dim.customWidth:
            widths[column_index_from_string(letter) - 1] = float(dim.width)
    if not widths:
        return []
    return [widths.get(index, 0.0) for index in range(max(widths) + 1)]


def _cell(cell: Any, report: LossReport, location: str) -> SheetCell:
    formula = ""
    value: object = cell.value
    if cell.data_type == "f" and isinstance(cell.value, str):
        # The FILE said formula. Never inferred from a leading "=" — a string cell whose
        # text starts with "=" is a label, and treating it as an expression is the defect
        # this parser exists to close.
        formula = cell.value
        value = None
        report.add(
            "formula_cached_value",
            f"{formula} — the last computed result is not carried; "
            "a reader recalculates it on open",
            location=location,
        )
    else:
        value = _json_scalar(value, report, location)
    align = str(getattr(cell.alignment, "horizontal", "") or "")
    if align and align not in _MODEL_ALIGNMENTS:
        report.add(
            "cell_style", f"horizontal alignment {align!r} has no model field", location=location
        )
        align = ""
    _report_cell_style(cell, report, location)
    return SheetCell(
        value=value,
        formula=formula,
        number_format=_number_format(cell),
        bold=bool(getattr(cell.font, "bold", False)),
        italic=bool(getattr(cell.font, "italic", False)),
        font_color=_rgb(getattr(cell.font, "color", None)),
        fill=_solid_fill(cell, report, location),
        align=align,
    )


def _json_scalar(value: object, report: LossReport, location: str) -> object:
    """Reduce a cell literal to something the model and the wire can both hold.

    openpyxl converts a date-formatted number into a real ``datetime``/``date``/``time``,
    and the model's ``value`` crosses the wire as JSON — where there is no date type. So a
    date arrives here as a Python object that :func:`sheet_json._value` would refuse, which
    would make a spreadsheet containing a single date unsaveable.

    It becomes its ISO 8601 string, and that degradation is REPORTED: a re-render writes
    text, so Excel will show the characters rather than a date it can sort or subtract.
    The alternative — writing an ISO-looking string back as a date — is the same sniffing
    that :class:`~personalclaw.documents.model.SheetCell` exists to abolish, just with a
    different pattern to match, so it is refused on principle rather than on effort.

    ``Decimal`` is the other openpyxl escapee and is a plain widening to ``float``, with no
    user-visible degradation to report.
    """
    import datetime as _dt
    from decimal import Decimal

    if isinstance(value, _dt.timedelta):
        report.add(
            "date_value",
            f"a duration of {value} is written back as text, not a duration",
            location=location,
        )
        return str(value)
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        report.add(
            "date_value",
            f"the date/time {value.isoformat()} is written back as text — a reader will "
            "no longer sort or calculate with it",
            location=location,
        )
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _number_format(cell: Any) -> str:
    """``"General"`` is openpyxl's word for "unformatted", which the model spells ``""``.

    Carrying the literal ``"General"`` would make every cell in every sheet declare a
    format, and a writer would then set one on all of them — noise in the file and a
    diff on every save.
    """
    fmt = str(cell.number_format or "")
    return "" if fmt == "General" else fmt


def _rgb(color: Any) -> str:
    """An openpyxl colour → ``RRGGBB``, or ``""`` when it is not a plain RGB one.

    ``color.rgb`` is a descriptor whose unset reading is not a colour at all (openpyxl
    hands back its own type-error prose), so the hex check is load-bearing rather than
    defensive. Stored values carry an alpha byte (``00FF3300``); the model holds
    ``RRGGBB``, so the leading pair is dropped — and a theme or indexed colour, which has
    no RGB, reads as unset.

    ``00000000`` is openpyxl's "nothing was set", NOT black: a deliberately black font is
    ``FF000000``. Reading the zero-alpha form as black would put an explicit colour on
    every cell of every sheet and write one back on every save.
    """
    raw = getattr(color, "rgb", None)
    if not isinstance(raw, str) or raw == "00000000":
        return ""
    hexed = raw.upper()
    if len(hexed) == 8:
        hexed = hexed[2:]
    if len(hexed) != 6 or any(ch not in "0123456789ABCDEF" for ch in hexed):
        return ""
    return hexed


def _solid_fill(cell: Any, report: LossReport, location: str) -> str:
    fill = getattr(cell, "fill", None)
    kind = str(getattr(fill, "fill_type", "") or "")
    if not kind:
        return ""
    if kind != "solid":
        report.add("cell_style", f"{kind} fill pattern has no model field", location=location)
        return ""
    return _rgb(getattr(fill, "start_color", None))


def _report_cell_style(cell: Any, report: LossReport, location: str) -> None:
    """Font and border attributes the model has no field for."""
    font = getattr(cell, "font", None)
    named = str(getattr(font, "name", "") or "")
    if named and named not in ("Calibri", "Aptos Narrow"):  # openpyxl/Excel defaults
        report.add("cell_style", f"font {named!r} has no model field", location=location)
    if getattr(font, "underline", None):
        report.add("cell_style", "underline has no model field", location=location)
    border = getattr(cell, "border", None)
    if border is not None and any(
        getattr(getattr(border, side, None), "style", None)
        for side in ("left", "right", "top", "bottom")
    ):
        report.add("cell_style", "cell borders have no model field", location=location)
    if str(getattr(cell.alignment, "vertical", "") or ""):
        report.add("cell_style", "vertical alignment has no model field", location=location)
