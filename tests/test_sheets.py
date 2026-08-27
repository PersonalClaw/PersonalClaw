"""DFE-7 — the spreadsheet round trip: a formula stays a formula.

**Every assertion here re-parses the WRITTEN BYTES.** A test that renders a model and
compares it to itself proves only that the writer is self-consistent; the claim under test
is that the *file* carries the distinction, so each case goes model → bytes →
:func:`parse_xlsx` → model, and several go one step further and read the bytes with
openpyxl directly so the evidence does not depend on our own parser either.

The defect this atom closes: before it, ``SheetModel`` held plain values, so a formula
could only be a string that happened to start with ``=``. openpyxl SNIFFS such a string
into a formula cell — which means the model was wrong in both directions at once. A
formula the caller meant survived by accident, and a *label* like ``"=TBD"`` was silently
turned into a formula that Excel opens as ``#NAME?``. Both directions are pinned below.
"""

from __future__ import annotations

import io

import pytest

from personalclaw.documents.model import Sheet, SheetCell, SheetModel
from personalclaw.documents.model_codec import MODEL_KINDS, get_codec
from personalclaw.documents.sheet_json import sheet_from_dict, sheet_to_dict
from personalclaw.documents.writers.xlsx_writer import render_xlsx
from personalclaw.documents.xlsx_parser import parse_xlsx


def _lap(model: SheetModel) -> SheetModel:
    """One full trip through the real writer AND the real parser."""
    parsed, _ = parse_xlsx(render_xlsx(model))
    return parsed


def _one(cell: SheetCell, **sheet_kwargs) -> SheetModel:
    return SheetModel(sheets=[Sheet(name="S", cells=[[cell]], **sheet_kwargs)])


def _cell_at(model: SheetModel, row: int = 0, col: int = 0) -> SheetCell:
    return model.sheets[0].cells[row][col]


def _openpyxl(data: bytes):
    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(data))


# ── clause 1: a formula stays a formula ──────────────────────────────────────


def test_a_formula_stays_a_formula_through_the_round_trip() -> None:
    """The atom's headline clause. ``=SUM(A1:A2)`` must come back as a FORMULA — in the
    ``formula`` field, not as a string sitting in ``value``."""
    model = SheetModel(
        sheets=[
            Sheet(
                name="Sales",
                cells=[
                    [SheetCell(value=10)],
                    [SheetCell(value=32)],
                    [SheetCell(formula="=SUM(A1:A2)")],
                ],
            )
        ]
    )

    back = _cell_at(_lap(model), row=2)

    assert back.formula == "=SUM(A1:A2)"
    # And it is NOT a string literal — the two fields are the whole distinction, so
    # asserting only `formula` would pass for a cell that carried both.
    assert back.value is None


def test_a_label_that_looks_like_a_formula_stays_a_label() -> None:
    """The same defect's other direction, and the one that corrupts a user's file.

    ``"=TBD"`` is a label somebody typed. openpyxl's own string sniffing turns it into a
    formula cell, which Excel then opens as ``#NAME?``. The writer must pin it back to a
    string, and the FILE must say ``s`` — read here with openpyxl rather than our parser,
    because the parser trusts ``data_type`` and would agree with itself.
    """
    data = render_xlsx(_one(SheetCell(value="=TBD")))

    assert _openpyxl(data)["S"]["A1"].data_type == "s"

    back = _cell_at(*(parse_xlsx(data)[0],))
    assert back.value == "=TBD"
    assert back.formula == ""


def test_a_formula_cell_is_written_as_a_formula_in_the_file() -> None:
    """The mirror of the test above: a declared formula must reach the file as ``f``.

    Together these two are the vacuity pair — each proves the other's guard is not simply
    forcing every cell to one type.
    """
    data = render_xlsx(_one(SheetCell(formula="=1+1")))

    assert _openpyxl(data)["S"]["A1"].data_type == "f"


def test_a_formula_must_declare_itself_with_a_leading_equals() -> None:
    """The model refuses a `formula` that is not one, so the field cannot become a second
    place a plain label hides."""
    with pytest.raises(ValueError, match="must start with '='"):
        SheetCell(formula="SUM(A1)")


def test_from_rows_leaves_a_formula_looking_value_a_literal() -> None:
    """The agent's convenience constructor supplies DATA. Promoting a value that looks
    like a formula would be exactly the guess the split exists to avoid."""
    model = SheetModel.from_rows({"S": [["=SUM(A1)"]]})

    assert model.sheets[0].cells[0][0].formula == ""
    assert model.sheets[0].cells[0][0].value == "=SUM(A1)"


# ── clause 2: a number format survives download/read-back ────────────────────


@pytest.mark.parametrize("code", ["0.0%", "#,##0.00", '#,##0.00" kg"', "0.00E+00"])
def test_a_number_format_survives_download_and_read_back(code: str) -> None:
    back = _cell_at(_lap(_one(SheetCell(value=0.25, number_format=code))))

    assert back.number_format == code
    assert back.value == 0.25  # the format must not have eaten the value


def test_a_date_format_survives_but_says_its_value_became_text() -> None:
    """A date format round-trips; the VALUE under it cannot, and that is reported.

    openpyxl turns a date-formatted number into a ``datetime``, and the model's value
    crosses the wire as JSON, which has no date type. It becomes an ISO string — so a
    re-render writes text where a date was, and a reader can no longer sort or subtract it.
    Writing an ISO-looking string back as a date would be the same sniffing this atom
    abolished, so the answer is an honest report item, not a cleverer guess.
    """
    back, loss = parse_xlsx(render_xlsx(_one(SheetCell(value=0.25, number_format="yyyy-mm-dd"))))

    assert _cell_at(back).number_format == "yyyy-mm-dd"
    assert _cell_at(back).value == "06:00:00"
    assert [item.kind for item in loss.items] == ["date_value"]
    assert loss.items[0].where == "S!A1"


def test_an_unformatted_cell_reads_back_unformatted() -> None:
    """The vacuity leg for the format assertions: ``""`` in, ``""`` out.

    Without it, a parser that returned openpyxl's literal ``"General"`` — or one that
    echoed the format of whatever cell it saw last — would pass every case above.
    """
    back = _cell_at(_lap(_one(SheetCell(value=0.25))))

    assert back.number_format == ""


# ── clause 3: a cell edit survives download/read-back ────────────────────────


def test_a_cell_edit_survives_download_and_read_back() -> None:
    """The editor's actual circuit: load a sheet, change ONE cell, save, re-read.

    Driven through the JSON codec both ways, because that is what the grid editor does —
    a test that mutated the dataclass in memory would skip the boundary where a dropped
    field actually happens.
    """
    original = render_xlsx(
        SheetModel(
            sheets=[
                Sheet(
                    name="Q1",
                    cells=[
                        [SheetCell(value="Region", bold=True), SheetCell(value="Rev", bold=True)],
                        [SheetCell(value="EMEA"), SheetCell(value=120, number_format="#,##0.00")],
                    ],
                )
            ]
        )
    )
    loaded, _ = parse_xlsx(original)

    payload = sheet_to_dict(loaded)
    payload["sheets"][0]["cells"][1][0]["value"] = "APAC"
    edited, _ = parse_xlsx(render_xlsx(sheet_from_dict(payload)))

    assert _cell_at(edited, row=1, col=0).value == "APAC"
    # The edit must not have disturbed its neighbours — a save that rewrites one cell and
    # flattens the rest is the fidelity failure, not a fix.
    assert _cell_at(edited, row=1, col=1).value == 120
    assert _cell_at(edited, row=1, col=1).number_format == "#,##0.00"
    assert _cell_at(edited, row=0, col=0).bold is True
    assert _cell_at(edited, row=0, col=0).value == "Region"


def test_an_edited_formula_is_still_a_formula_after_the_save() -> None:
    """Editing a formula through the codec must not degrade it into text — the grid edits
    formulas AS formulas, which is the clause's whole point."""
    loaded, _ = parse_xlsx(render_xlsx(_one(SheetCell(formula="=1+1"))))

    payload = sheet_to_dict(loaded)
    payload["sheets"][0]["cells"][0][0]["formula"] = "=2+2"
    data = render_xlsx(sheet_from_dict(payload))

    assert _openpyxl(data)["S"]["A1"].data_type == "f"
    assert _cell_at(parse_xlsx(data)[0]).formula == "=2+2"


# ── the rest of the styled model ─────────────────────────────────────────────


def test_cell_presentation_survives_the_round_trip() -> None:
    cell = SheetCell(
        value="Total",
        bold=True,
        italic=True,
        font_color="FF3300",
        fill="FFEEAA",
        align="center",
    )

    back = _cell_at(_lap(_one(cell)))

    assert (back.bold, back.italic) == (True, True)
    assert back.font_color == "FF3300"
    assert back.fill == "FFEEAA"
    assert back.align == "center"


def test_a_plain_cell_reads_back_with_no_presentation() -> None:
    """Vacuity for the case above: an unstyled cell must not acquire a style.

    ``00000000`` is openpyxl's "unset", not black — reading it as a colour would put an
    explicit font colour on every cell of every sheet and write one back on every save.
    """
    back = _cell_at(_lap(_one(SheetCell(value="plain"))))

    assert (back.bold, back.italic) == (False, False)
    assert (back.font_color, back.fill, back.align) == ("", "", "")


def test_column_widths_and_merges_and_a_frozen_header_survive() -> None:
    model = SheetModel(
        sheets=[
            Sheet(
                name="Wide",
                cells=[[SheetCell(value="a"), SheetCell(value="b")]],
                column_widths=[0.0, 22.5],
                merges=["A3:B3"],
                frozen_header=True,
            )
        ]
    )

    back = _lap(model).sheets[0]

    assert back.column_widths == [0.0, 22.5]  # dense and index-aligned; 0 = default
    assert back.merges == ["A3:B3"]
    assert back.frozen_header is True


def test_a_sheet_with_no_geometry_reads_back_with_none() -> None:
    """Vacuity for the geometry assertions."""
    back = _lap(_one(SheetCell(value="x"))).sheets[0]

    assert back.column_widths == []
    assert back.merges == []
    assert back.frozen_header is False


def test_cell_types_are_preserved_including_bool_distinctly_from_int() -> None:
    """``bool`` IS an ``int`` in Python, so an unordered check writes True as 1."""
    model = SheetModel(
        sheets=[
            Sheet(
                name="T",
                cells=[
                    [
                        SheetCell(value=True),
                        SheetCell(value=1),
                        SheetCell(value=1.5),
                        SheetCell(value="1"),
                    ]
                ],
            )
        ]
    )

    row = _lap(model).sheets[0].cells[0]

    assert row[0].value is True
    assert [type(c.value).__name__ for c in row] == ["bool", "int", "float", "str"]


def test_a_trailing_empty_cell_is_not_invented_on_read_back() -> None:
    """An empty cell is not stored in the file, so a trailing one does not come back — and
    the parser must not pad the row to a width the file never had. True of every
    spreadsheet, and worth pinning because the alternative (guessing a width) would make
    the round trip unstable."""
    model = SheetModel(
        sheets=[Sheet(name="S", cells=[[SheetCell(value="a"), SheetCell(value=None)]])]
    )

    assert [c.value for c in _lap(model).sheets[0].cells[0]] == ["a"]


def test_sheet_order_is_preserved() -> None:
    """Sheet order is part of the document, which is why the model holds a list."""
    model = SheetModel.from_rows({"Zeta": [["z"]], "Alpha": [["a"]], "Mid": [["m"]]})

    assert [s.name for s in _lap(model).sheets] == ["Zeta", "Alpha", "Mid"]


def test_parse_write_parse_is_stable() -> None:
    """A second lap must change nothing. An unstable round trip means every save mutates
    the document a little, which is the silent corruption this plan exists to prevent."""
    first = _lap(
        SheetModel(
            sheets=[
                Sheet(
                    name="Books",
                    cells=[
                        [SheetCell(value="Item", bold=True), SheetCell(value="Qty", bold=True)],
                        [SheetCell(value="Pens"), SheetCell(value=12)],
                        [SheetCell(value="Total"), SheetCell(formula="=SUM(B2:B2)")],
                    ],
                    column_widths=[18.0],
                    frozen_header=True,
                )
            ]
        )
    )

    assert _lap(first) == first


def test_rows_is_derived_and_shows_the_formula_not_a_stale_value() -> None:
    """``Sheet.rows`` is a property, so it cannot disagree with the cells it comes from —
    the reason the wire carries cells only."""
    sheet = Sheet(
        name="S",
        cells=[[SheetCell(value="x"), SheetCell(formula="=A1", value="stale cache")]],
    )

    assert sheet.rows == [["x", "=A1"]]


# ── the loss report ──────────────────────────────────────────────────────────


def test_a_formula_reports_that_its_cached_result_is_not_carried() -> None:
    """Honest, not silent: the parse keeps formulas rather than cached values, so the last
    computed result is genuinely gone until a reader recalculates."""
    _, loss = parse_xlsx(render_xlsx(_one(SheetCell(formula="=1+1"))))

    assert loss.lossless is False
    assert [item.kind for item in loss.items] == ["formula_cached_value"]
    # Located in the SHEET's own terms. "block 0" would be a lie about a spreadsheet.
    assert loss.items[0].where == "S!A1"


def test_a_plain_sheet_is_lossless() -> None:
    """Vacuity for every loss assertion here: without it a parser that reported an item
    per cell would satisfy all of them."""
    _, loss = parse_xlsx(render_xlsx(SheetModel.from_rows({"S": [["a", 1]]})))

    assert loss.lossless is True
    assert loss.summary() == "no losses"


def test_a_sheet_feature_the_model_cannot_hold_is_reported_not_dropped() -> None:
    """Conditional formatting has no model field, so a re-render drops it. It is named."""
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = 5
    ws.conditional_formatting.add("A1:A9", CellIsRule(operator="lessThan", formula=["3"]))
    buf = io.BytesIO()
    wb.save(buf)

    _, loss = parse_xlsx(buf.getvalue())

    assert "sheet_feature" in loss.kinds()
    assert any("conditional formatting" in item.detail for item in loss.of_kind("sheet_feature"))


def test_a_cell_style_with_no_model_field_is_reported() -> None:
    """A border is real formatting a save would drop, so it is reported rather than
    silently lost — and located at the cell."""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["B2"] = "boxed"
    ws["B2"].border = Border(left=Side(style="thin"))
    buf = io.BytesIO()
    wb.save(buf)

    _, loss = parse_xlsx(buf.getvalue())

    borders = [i for i in loss.of_kind("cell_style") if "border" in i.detail]
    assert borders and borders[0].where == "S!B2"


def test_an_explicit_row_height_is_reported_per_row_and_located_at_that_row() -> None:
    """A row the author resized is real layout the model cannot hold, so a re-render
    silently returns it to the default. It is named once per resized row.

    Row 2 carries no height and is this case's own vacuity leg: a parser that reported
    every row it walked — or one item for the sheet — would fail the count, not just the
    location. ``test_a_plain_sheet_is_lossless`` covers the other direction.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "tall"
    ws["A2"] = "default"
    ws.row_dimensions[1].height = 42
    buf = io.BytesIO()
    wb.save(buf)

    _, loss = parse_xlsx(buf.getvalue())

    heights = loss.of_kind("row_height")
    # Located in the SHEET's own terms, one item, for the ONE row that set a height.
    assert [item.where for item in heights] == ["S!1"]
    assert "42" in heights[0].detail


def test_a_sheet_whose_rows_were_never_resized_reports_no_row_height() -> None:
    """The explicit vacuity leg for the row-height assertion: two rows of real content,
    neither resized, and the kind must not appear at all. Without it, a parser that
    emitted ``row_height`` for every row present would satisfy the case above."""
    _, loss = parse_xlsx(render_xlsx(SheetModel.from_rows({"S": [["a", 1], ["b", 2]]})))

    assert "row_height" not in loss.kinds()


# ── the JSON boundary ────────────────────────────────────────────────────────


def test_the_json_codec_round_trips_the_model_unchanged() -> None:
    model = SheetModel(
        sheets=[
            Sheet(
                name="S",
                cells=[[SheetCell(value=1, number_format="0.0%", bold=True)]],
                column_widths=[12.0],
                merges=["A1:B1"],
                frozen_header=True,
            )
        ]
    )

    assert sheet_from_dict(sheet_to_dict(model)) == model


def test_the_wire_shape_carries_cells_only_never_a_second_row_view() -> None:
    """``rows`` is derived, so it must not appear on the wire — a shipped copy of the
    cells is a second representation that goes stale on the first edit."""
    payload = sheet_to_dict(SheetModel.from_rows({"S": [["a"]]}))

    assert "rows" not in payload["sheets"][0]
    assert payload["sheets"][0]["cells"][0][0]["value"] == "a"


def test_an_unknown_field_is_refused_rather_than_dropped() -> None:
    """Strict, for ``model_json``'s reason: a client that misspells ``number_format`` must
    learn its formatting was not saved, not find out in Excel."""
    payload = sheet_to_dict(SheetModel.from_rows({"S": [["a"]]}))
    payload["sheets"][0]["cells"][0][0]["numberformat"] = "0.0%"

    with pytest.raises(ValueError, match="numberformat"):
        sheet_from_dict(payload)


def test_a_composite_cell_value_is_refused() -> None:
    """A list would reach the writer as ``str()`` of a Python repr, which is not what any
    user typed."""
    payload = sheet_to_dict(SheetModel.from_rows({"S": [["a"]]}))
    payload["sheets"][0]["cells"][0][0]["value"] = ["a", "b"]

    with pytest.raises(ValueError, match="must be a string, number, boolean or null"):
        sheet_from_dict(payload)


# ── the codec table ──────────────────────────────────────────────────────────


def test_every_declared_model_kind_resolves_to_a_codec() -> None:
    """``MODEL_KINDS`` is declared rather than computed, so this is what stops it naming a
    kind with no runtime behind it — the route answers a capability question with it."""
    assert MODEL_KINDS
    for kind in MODEL_KINDS:
        codec = get_codec(kind)
        assert codec is not None, f"{kind} is advertised with no codec"
        assert codec.kind == kind


def test_a_kind_with_no_codec_resolves_to_none() -> None:
    """Vacuity for the test above: ``get_codec`` must be capable of saying no, or the
    loop would pass for any list of strings at all."""
    assert get_codec("pptx") is None
    assert get_codec("png") is None
    assert "pptx" not in MODEL_KINDS


def test_the_xlsx_codec_is_wired_to_the_shipped_parser_and_writer() -> None:
    """The call site, not the mechanism: the codec the ROUTE looks up must be the parser
    and serializer this suite proved, not a second pair."""
    codec = get_codec("xlsx")
    assert codec is not None
    assert codec.parse is parse_xlsx
    assert codec.to_dict is sheet_to_dict
    assert codec.from_dict is sheet_from_dict
