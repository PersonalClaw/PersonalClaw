"""Document generation — the writer seam, the markup parsers, and round-trip validity.

The central question for generated files is VALIDITY, and a subtly-invalid .docx that our
own reader accepts but Word rejects is the worst outcome. So the strongest proof available
in-process is a round trip through `knowledge/readers.py` — the same code that ingests
user-uploaded documents. That is necessary but NOT sufficient: the plan's V-task requires
opening the output in a real application before the session may close.
"""

from __future__ import annotations

import pytest

from personalclaw.documents import available_formats, get_writer
from personalclaw.documents.from_markup import (
    deck_from_markdown,
    document_from_html,
    document_from_markdown,
)
from personalclaw.documents.model import Block, DeckModel, DocumentModel, SheetModel
from personalclaw.knowledge.readers import FileReader


def _write(tmp_path, fmt: str, model, name="out"):
    data = get_writer(fmt)(model)
    path = tmp_path / f"{name}.{fmt}"
    path.write_bytes(data)
    return path, data


# ── the registry ──────────────────────────────────────────────────────────────


def test_registry_reports_only_usable_formats():
    """Registration IS the availability check — a writer whose library is missing never
    registers, so this can't advertise a format that would fail on use."""
    assert set(available_formats()) >= {"docx", "xlsx"}


def test_an_unknown_format_returns_none_rather_than_raising():
    """A caller turns this into a typed refusal; raising would force a try/except at
    every call site."""
    assert get_writer("dwg") is None
    assert get_writer("") is None


def test_format_lookup_is_case_insensitive():
    assert get_writer("DOCX") is not None


def test_a_writer_rejects_the_wrong_model_type():
    with pytest.raises(TypeError):
        get_writer("docx")(SheetModel(sheets={"a": []}))
    with pytest.raises(TypeError):
        get_writer("xlsx")(DocumentModel(title="x"))


def test_block_rejects_an_unknown_kind():
    """A writer must handle every declared kind; letting an unknown one through would
    mean silently dropping the user's content."""
    with pytest.raises(ValueError, match="unknown block kind"):
        Block(kind="carousel")


def test_block_clamps_the_heading_level():
    assert Block(kind="heading", level=99).level == 6
    assert Block(kind="heading", level=0).level == 1


# ── markdown → model ──────────────────────────────────────────────────────────


def test_markdown_parses_every_shape_a_generated_document_uses():
    md = """# Title

Opening paragraph.

## Section

- one
- two

1. first
2. second

| A | B |
|---|---|
| 1 | 2 |

```python
x = 1
```

---

Closing.
"""
    model = document_from_markdown(md)
    kinds = [b.kind for b in model.blocks]

    assert model.title == "Title"  # a leading H1 becomes the title, not a dup heading
    assert kinds == [
        "paragraph",
        "heading",
        "bullets",
        "numbered",
        "table",
        "code",
        "pagebreak",
        "paragraph",
    ]
    table = next(b for b in model.blocks if b.kind == "table")
    assert table.rows == [["A", "B"], ["1", "2"]]  # the |---| separator is not a row
    code = next(b for b in model.blocks if b.kind == "code")
    assert code.text == "x = 1"


def test_inline_emphasis_is_stripped_but_code_fences_are_verbatim():
    model = document_from_markdown(
        "Some **bold** and `mono` and [a link](http://x).\n\n```\nkeep **this**\n```\n"
    )
    para = next(b for b in model.blocks if b.kind == "paragraph")
    assert para.text == "Some bold and mono and a link."
    code = next(b for b in model.blocks if b.kind == "code")
    assert code.text == "keep **this**", "a fence must not be reinterpreted"


def test_an_explicit_title_keeps_a_leading_h1_as_a_heading():
    model = document_from_markdown("# Not the title\n\nbody\n", title="Given")
    assert model.title == "Given"
    assert model.blocks[0].kind == "heading"


def test_unrecognized_lines_become_paragraphs_rather_than_vanishing():
    model = document_from_markdown("> a quote\n\n:::weird:::\n")
    assert [b.kind for b in model.blocks] == ["paragraph", "paragraph"]


def test_an_unterminated_code_fence_still_keeps_its_content():
    model = document_from_markdown("```\nlost?\n")
    assert any(b.kind == "code" and "lost?" in b.text for b in model.blocks)


def test_empty_markdown_yields_an_empty_but_valid_model():
    model = document_from_markdown("")
    assert model.blocks == [] and model.title == ""


def test_html_is_sanitized_and_credentials_redacted():
    """HTML here is agent- or web-authored, so untrusted. Routed through the platform's
    EXISTING sanitizer + credential redactor, never a second implementation."""
    model = document_from_html(
        "<h1>Doc</h1><script>steal()</script><p>key AKIAIOSFODNN7EXAMPLE here</p>"
    )
    body = " ".join(b.text for b in model.blocks) + model.title

    assert "steal()" not in body
    assert "<script" not in body
    assert "AKIAIOSFODNN7EXAMPLE" not in body, "a credential must not survive into a file"


# ── round trip: our own readers must read what we write ───────────────────────


def test_a_generated_docx_re_reads_through_the_real_reader(tmp_path):
    md = "# Quarterly\n\nRevenue grew.\n\n## Details\n\n- EMEA up\n- APAC flat\n"
    path, data = _write(tmp_path, "docx", document_from_markdown(md))

    assert len(data) > 1000  # a real OOXML package, not an empty stub
    text, meta = FileReader().read(str(path))

    assert meta["format"] == "docx"
    assert "Quarterly" in text
    assert "## Details" in text, "heading level must survive the round trip"
    assert "EMEA up" in text and "APAC flat" in text


def test_a_generated_docx_table_survives_the_round_trip(tmp_path):
    """Tables were previously dropped by the docx READER (it walked only
    `doc.paragraphs`). Fixed in this change — so this asserts the whole path."""
    md = "# T\n\n| Region | Q1 |\n|---|---|\n| EMEA | 120 |\n"
    path, _ = _write(tmp_path, "docx", document_from_markdown(md))

    text, meta = FileReader().read(str(path))

    assert meta["table_count"] == 1
    assert "| Region | Q1 |" in text
    assert "| EMEA | 120 |" in text


def test_a_generated_xlsx_re_reads_with_numbers_still_numeric(tmp_path):
    """A spreadsheet whose numbers arrived as text can't be summed, which defeats the
    point of generating one."""
    from openpyxl import load_workbook

    model = SheetModel(sheets={"Sales": [["Region", "Q1"], ["EMEA", 120], ["APAC", 99.5]]})
    path, _ = _write(tmp_path, "xlsx", model)

    text, meta = FileReader().read(str(path))
    assert meta["format"] == "xlsx"
    assert "| EMEA | 120 |" in text

    ws = load_workbook(path)["Sales"]
    assert [type(c.value).__name__ for c in ws[2]] == ["str", "int"]
    assert type(ws[3][1].value).__name__ == "float"


def test_xlsx_preserves_bool_distinctly_from_int(tmp_path):
    """`bool` IS an `int` in Python, so an unordered isinstance check would write True
    as 1 and lose the distinction the model preserves deliberately."""
    from openpyxl import load_workbook

    path, _ = _write(tmp_path, "xlsx", SheetModel(sheets={"S": [["flag"], [True]]}))
    assert load_workbook(path)["S"]["A2"].value is True


def test_xlsx_sanitizes_illegal_sheet_names_and_dedupes(tmp_path):
    """Excel refuses some names outright; a rejected name would fail the whole write."""
    from openpyxl import load_workbook

    model = SheetModel(sheets={"a/b:c*d?e[f]": [["x"]], "x" * 40: [["y"]]})
    path, _ = _write(tmp_path, "xlsx", model)

    names = load_workbook(path).sheetnames
    assert all(not (set(n) & set(r"[]:*?/\\")) for n in names)
    assert all(len(n) <= 31 for n in names)


def test_an_empty_sheet_model_still_produces_a_valid_workbook(tmp_path):
    from openpyxl import load_workbook

    path, _ = _write(tmp_path, "xlsx", SheetModel(sheets={}))
    assert load_workbook(path).sheetnames  # a workbook with zero sheets is invalid


def test_ragged_table_rows_are_normalized_not_truncated(tmp_path):
    """python-docx needs a fixed column count; truncating would silently lose cells."""
    model = DocumentModel(blocks=[Block(kind="table", rows=[["a", "b", "c"], ["1"]])])
    path, _ = _write(tmp_path, "docx", model)

    text, _ = FileReader().read(str(path))
    assert "| a | b | c |" in text


def test_an_image_block_renders_a_placeholder_rather_than_vanishing(tmp_path):
    """Resolving an artifact reference to bytes is the caller's job (it owns the store).
    Dropping the block would lose the fact that an image belonged here."""
    model = DocumentModel(blocks=[Block(kind="image", artifact_slug="sales-chart")])
    path, _ = _write(tmp_path, "docx", model)

    text, _ = FileReader().read(str(path))
    assert "sales-chart" in text


# ── deck outline ──────────────────────────────────────────────────────────────


def test_deck_from_markdown_splits_slides_and_captures_notes():
    md = """# The Deck

## First

- point one
- point two

<!-- notes: say hello -->

## Second

body line
"""
    deck = deck_from_markdown(md)

    assert deck.title == "The Deck"
    assert [s.title for s in deck.slides] == ["First", "Second"]
    assert deck.slides[0].body == ["point one", "point two"]
    assert deck.slides[0].notes == "say hello"
    assert deck.slides[1].body == ["body line"]


def test_deck_body_before_any_heading_gets_an_opening_slide():
    deck = deck_from_markdown("orphan line\n")
    assert len(deck.slides) == 1 and deck.slides[0].body == ["orphan line"]


def test_deck_notes_are_kept_out_of_the_visible_body():
    deck = deck_from_markdown("## S\n\n<!-- notes: hidden -->\n- shown\n")
    assert deck.slides[0].body == ["shown"]
    assert "hidden" not in " ".join(deck.slides[0].body)


def test_empty_deck_markdown_is_a_valid_empty_deck():
    assert deck_from_markdown("").slides == []
    assert isinstance(deck_from_markdown(""), DeckModel)


# ── artifact kinds + the coercion hardening ───────────────────────────────────


def test_the_new_document_kinds_are_registered_in_both_sets():
    """A binary kind must be in ALLOWED_KINDS *and* BINARY_KINDS. Being in neither is
    exactly how generated video ended up stored as an image (issue #94)."""
    from personalclaw.artifacts.models import ALLOWED_KINDS, BINARY_KINDS

    for kind in ("docx", "xlsx", "pdf", "video"):
        assert kind in ALLOWED_KINDS, kind
        assert kind in BINARY_KINDS, kind
    # csv is a TEXT kind: it round-trips as text and needs no binary body.
    assert "csv" in ALLOWED_KINDS and "csv" not in BINARY_KINDS


def test_every_binary_kind_has_a_mime_extension_mapping():
    """The raw endpoint derives Content-Type from the stored extension — an unmapped
    mime serves a download the OS can't open."""
    from personalclaw.artifacts.models import _MIME_TO_EXT, BINARY_KINDS

    exts = set(_MIME_TO_EXT.values())
    for kind in BINARY_KINDS:
        if kind == "image":  # image maps via several mimes, none named "image"
            continue
        assert kind in exts or kind in {"video"}, f"no mime→ext mapping produces {kind}"


def test_create_binary_raises_on_a_non_binary_kind(tmp_path):
    """This used to coerce silently to "image" — the #94 bug class. A programming error
    must fail loudly so a newly-added kind can't be quietly mis-stored."""
    from personalclaw.artifacts.native import NativeArtifactProvider

    prov = NativeArtifactProvider(root=tmp_path)
    with pytest.raises(ValueError, match="non-binary kind"):
        prov.create_binary(name="X", data=b"x", mime="text/plain", kind="markdown")


def test_create_binary_accepts_a_registered_document_kind(tmp_path):
    from personalclaw.artifacts.native import NativeArtifactProvider

    prov = NativeArtifactProvider(root=tmp_path)
    art = prov.create_binary(
        name="Doc",
        data=b"PK\x03\x04stub",
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        kind="docx",
    )
    assert art.kind == "docx", "the kind must survive, not be coerced"
    assert art.content.startswith("/api/artifacts/"), "content is a raw ref, never bytes"


def test_a_generated_document_is_stored_under_its_real_kind(tmp_path):
    """End to end through the writer + store: the artifact keeps its format identity."""
    from personalclaw.artifacts.native import NativeArtifactProvider

    prov = NativeArtifactProvider(root=tmp_path)
    data = get_writer("docx")(document_from_markdown("# T\n\nbody\n"))
    art = prov.create_binary(
        name="T",
        data=data,
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        kind="docx",
    )
    stored, mime = prov.raw_bytes(art.slug)
    assert stored == data, "the bytes must round-trip unchanged"
    assert mime.endswith("wordprocessingml.document")


# ── S2: pptx ──────────────────────────────────────────────────────────────────


def test_pptx_is_available_and_round_trips_titles_bodies_and_notes(tmp_path):
    """The reader identifies a slide title via `slide.shapes.title`, so a deck built from
    free-floating text boxes would round-trip with every title lost."""
    md = """# The Deck

## First slide

- alpha
- beta

<!-- notes: mention the numbers -->

## Second slide

- gamma
"""
    assert "pptx" in available_formats()
    path, data = _write(tmp_path, "pptx", deck_from_markdown(md))

    assert len(data) > 5000
    text, meta = FileReader().read(str(path))

    assert meta["format"] == "pptx"
    assert meta["slide_count"] == 3  # title slide + two content slides
    assert "First slide" in text and "Second slide" in text
    assert "alpha" in text and "gamma" in text
    assert "mention the numbers" in text, "speaker notes must survive"


def test_pptx_body_placeholder_is_found_by_index_not_identity(tmp_path):
    """python-pptx returns a NEW proxy on each `shapes.title` access, so
    `shape is slide.shapes.title` is False even for the title placeholder. An identity
    check made the first body line overwrite the title — measured, not assumed."""
    from personalclaw.documents.model import DeckModel, Slide

    deck = DeckModel(slides=[Slide(title="Real Title", body=["first bullet"])])
    path, _ = _write(tmp_path, "pptx", deck)

    text, _ = FileReader().read(str(path))
    assert "Slide 1: Real Title" in text, "the title must not be overwritten by the body"
    assert "first bullet" in text


def test_a_slide_with_no_body_still_renders(tmp_path):
    from personalclaw.documents.model import DeckModel, Slide

    path, _ = _write(tmp_path, "pptx", DeckModel(slides=[Slide(title="Only a title")]))
    text, _ = FileReader().read(str(path))
    assert "Only a title" in text


def test_a_deck_image_reference_is_recorded_in_the_notes(tmp_path):
    """Resolving an artifact to bytes is the caller's job; dropping the block would lose
    the fact that an image belonged on the slide."""
    from personalclaw.documents.model import DeckModel, Slide

    deck = DeckModel(slides=[Slide(title="Chart", artifact_slug="sales-chart")])
    path, _ = _write(tmp_path, "pptx", deck)

    text, _ = FileReader().read(str(path))
    assert "sales-chart" in text


# ── S2: pdf ───────────────────────────────────────────────────────────────────


def test_pdf_is_unconditionally_available():
    """reportlab is a CORE dependency (owner ruling), so pdf is never a
    sometimes-present format — the agent can offer it on any install."""
    assert "pdf" in available_formats()


def test_a_generated_pdf_re_reads_through_the_existing_pdfplumber_reader(tmp_path):
    md = """# Q3 Review

Revenue grew across all regions.

## Highlights

- EMEA up
- APAC flat

1. Hire
2. Ship

| Region | Q1 |
|--------|----|
| EMEA   | 120|
"""
    path, data = _write(tmp_path, "pdf", document_from_markdown(md))

    assert data.startswith(b"%PDF"), "must be a real PDF, not a stub"
    text, meta = FileReader().read(str(path))

    assert meta["format"] == "pdf"
    assert "Q3 Review" in text and "Highlights" in text
    assert "EMEA up" in text and "APAC flat" in text
    assert "Hire" in text and "Ship" in text
    assert "120" in text, "table content must survive"


def test_pdf_bullets_extract_as_text_not_cid_garbage(tmp_path):
    """reportlab's default bullet is ZapfDingbats char 127, whose CID has no unicode
    mapping — every bullet extracted as the literal string "(cid:127)", corrupting the
    text of any generated PDF later ingested or searched."""
    path, _ = _write(tmp_path, "pdf", document_from_markdown("# T\n\n- alpha\n- beta\n"))

    text, _ = FileReader().read(str(path))

    assert "(cid:" not in text, "bullet glyph leaked an unmapped CID into the text"
    assert "alpha" in text and "beta" in text


def test_pdf_escapes_markup_characters_from_document_content(tmp_path):
    """Platypus parses mini-HTML inside Paragraph text, so a raw `<` or `&` from content
    would vanish or raise mid-build."""
    path, _ = _write(
        tmp_path, "pdf", document_from_markdown("# T\n\nUse <angle> & ampersand chars.\n")
    )

    text, _ = FileReader().read(str(path))
    assert "<angle>" in text and "&" in text


def test_an_empty_document_model_still_builds_every_format(tmp_path):
    """A zero-flowable reportlab build raises; the other writers must not produce a
    corrupt file either. An empty document is valid input."""
    for fmt in ("docx", "pdf"):
        _, data = _write(tmp_path, fmt, DocumentModel(), name=f"empty-{fmt}")
        assert data, fmt


# ── S2: the round trip (T2.4) ─────────────────────────────────────────────────


def test_exporting_a_text_artifact_as_a_document(tmp_path, monkeypatch):
    """The round trip: something already in the library comes back OUT as a real file,
    through the SAME writer path as a fresh generation — no parallel export pipeline."""
    from personalclaw.artifacts import registry
    from personalclaw.artifacts.native import NativeArtifactProvider
    from personalclaw.mcp_artifacts import _resolve_document_source

    prov = NativeArtifactProvider(root=tmp_path)
    monkeypatch.setitem(registry._providers, "native", prov)
    prov.create(
        name="Research Notes",
        content="# Research Notes\n\nFindings.\n\n- one\n",
        kind="markdown",
    )

    body, title = _resolve_document_source(prov, "research-notes")

    assert body is not None and "Findings." in body
    assert title == "Research Notes"


def test_exporting_a_binary_artifact_is_refused(tmp_path, monkeypatch):
    """A binary artifact's `content` is a raw URL, not text — exporting one would write
    the URL into the document body."""
    from personalclaw.artifacts import registry
    from personalclaw.artifacts.native import NativeArtifactProvider
    from personalclaw.mcp_artifacts import _resolve_document_source

    prov = NativeArtifactProvider(root=tmp_path)
    monkeypatch.setitem(registry._providers, "native", prov)
    art = prov.create_binary(name="Pic", data=b"\x89PNG", mime="image/png", kind="image")

    body, _ = _resolve_document_source(prov, art.slug)

    assert body is None, "a binary artifact must not be exported as document text"


def test_an_unknown_source_resolves_to_none(tmp_path, monkeypatch):
    from personalclaw.artifacts import registry
    from personalclaw.artifacts.native import NativeArtifactProvider
    from personalclaw.mcp_artifacts import _resolve_document_source

    prov = NativeArtifactProvider(root=tmp_path)
    monkeypatch.setitem(registry._providers, "native", prov)

    assert _resolve_document_source(prov, "nope-nope")[0] is None


def test_the_pptx_kind_is_registered_in_both_sets():
    from personalclaw.artifacts.models import ALLOWED_KINDS, BINARY_KINDS

    assert "pptx" in ALLOWED_KINDS and "pptx" in BINARY_KINDS
